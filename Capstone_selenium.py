from selenium import webdriver
#get, no need to wait the page loading 
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
desired_capabilities = DesiredCapabilities.CHROME
desired_capabilities["pageLoadStrategy"] = "none"
from selenium.webdriver.chrome.options import Options
from time import sleep  #to rest 
import re
import json
import openpyxl
import random

# [\000-\010]|[\013-\014]|[\016-\037]
class x_xlsx():
    def __init__(self,name,sheet='Sheet1'):
        self.name01=name
        self.sheet=sheet

    # read the sheet
    def readsheet(self, xuan01=''):
        workbook = openpyxl.load_workbook(self.name01)
        # sheet = wb.get_sheet_by_name(sheet_name) not recommend to use this method, might hv bug
        # n01=workbook.get_sheet_names()
        n01 = workbook.sheetnames
        print('sheetname：', n01)

        if xuan01 != '':
            sheet = workbook[n01[xuan01]]
        else:
            sheet = workbook[self.sheet]

        zhuan02 = []
        for row in sheet.rows:
            zhuan01 = []
            for cell in row:
                zhuan01.append(cell.value)
                # print(cell.value, "\t", end="")
            zhuan02.append(zhuan01)

        workbook.close()
        return zhuan02

    # build sheet
    def buildsheet(self,tou):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title =self.sheet
        # data entry starting from(1,1)
        # sheet.cell(1, 1, value=str('5'))
        sheet.append(tou)
        workbook.save(self.name01)
        print("創建成功。。。。。。。")

    # sheet edit，追加
    def sheetedit02(self,data_list):
        wb = openpyxl.load_workbook(self.name01)
        # change to target workbook
        ws = wb[self.sheet]
        # data entry starting from(1,1)
        # ws.cell(1,1,'ass')
        # ws.cell(row=1, column=1).fill = sty.PatternFill(fill_type='solid', fgColor="0d5330")
        # ws['a1']='uigufyc'
        zhuan01=[]
        for i01 in data_list:
            n01 = re.sub(re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]'), '', str(i01))

            if type(i01) == int or type(i01) == float:
                try:
                    n01 = int(n01)
                except:
                    try:
                        n01 = float(n01)
                    except:
                        pass

            zhuan01.append(n01)

        ws.append(zhuan01)
        wb.save(self.name01)
        print('data updated。。。。。。。。。。。。')


    # sheet edit，追加
    def sheetedit03(self,data_list):
        wb = openpyxl.load_workbook(self.name01)
        # change to target workbook
        ws = wb[self.sheet]
        for data in data_list:
            zhuan01=[]
            for i01 in data:
                n01=re.sub(re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]'),'',str(i01))

                if type(i01) == int or type(i01) == float:
                    try:
                        n01 = int(n01)
                    except:
                        try:
                            n01 = float(n01)
                        except:
                            pass

                zhuan01.append(n01)

            ws.append(zhuan01)

        wb.save(self.name01)
        print('数據集已更新。。。。。。。。。。。。')


options = Options()
options.add_experimental_option('excludeSwitches', ['enable-logging'])
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("window-size=1000,800")


driver=webdriver.Chrome(options=options)

xx01=['2018-01','2018-02','2018-03','2018-04','2018-05','2018-06','2018-07','2018-08','2018-09','2018-10','2018-11','2018-12',
'2019-01','2019-02','2019-03','2019-04','2019-05','2019-06','2019-07','2019-08','2019-09','2019-10','2019-11','2019-12',
'2020-01','2020-02','2020-03','2020-04','2020-05','2020-06','2020-07','2020-08','2020-09','2020-10','2020-11','2020-12',
'2021-01','2021-02','2021-03','2021-04','2021-05','2021-06','2021-07','2021-08','2021-09','2021-10','2021-11','2021-12']

yue01=x_xlsx('result.xlsx')
yue01.buildsheet([])


url_id01='235581265'
# url_id01='234243008'
# url_id01='231566248'

y02=0
while 1:
    y02=y02+1

    print('*'*120)
    print(f'page{y02}。。。。。。。')

    url01='https://www.dcard.tw/service/api/v2/forums/hkmaculife/posts?limit=100&before={}'.format(url_id01)
    print('website：',url01)


    a01=0
    while 1:
        a01=a01+1

        driver.get(url01)
        sleep(2.4)

        cha01=driver.find_elements_by_css_selector('[id="cf-wrapper"]')
        if cha01:
            print('出現驗證碼***********')

            if a01<3:
                sleep(30)
            else:
                sleep(90)   #could input by myself, or wait it 30s 
        else:
            break

    a01=0
    while 1:
        try:

            a01=a01+1
            if a01>6:
                print('waiting to reopen the page。。。。。。。。。。。')
                sleep(60)

                driver.get(url01)
                sleep(2.4)

            data01=json.loads(re.sub('<.*?>','',driver.page_source,flags=re.DOTALL))
            print('詳情個數：',len(data01))

            break
        except Exception as e:
            print('data error：',e)

        sleep(1.2)

    biao01=0
    nr_list=[]
    for y01 in range(0,len(data01)):
        n01=data01[y01]

        n02=str(n01['id'])
        url_id01 =n02

        n03=n01['title']
        n04=n01['excerpt']
        n05=n01['createdAt']
        n06=n01['commentCount']
        n07=n01['likeCount']
        n08=n01['topics']
        n09=n01['gender']
        n010=n01['school']
        n011=n01['reactions']

        sj01=n05.split('-')
        sj001='-'.join(sj01[:2])

        if y01==0:
            print('time:',n05)

        if n05.find('2017-12')!=-1:
            biao01=1
            print('finish scraping 01.。。。。')


        if sj001 in xx01:
            print('match the data。。。。')
        else:
            continue

        zhuan01=[]
        for i01 in range(2,12):
            zhuan01.append(eval('n0{}'.format(i01)))
            # print(i01,eval('n0{}'.format(i01)))

        print(y01,zhuan01)
        nr_list.append(zhuan01)

        

    if nr_list:
        print('enter the data:',len(nr_list))
        yue01.sheetedit03(nr_list)

    sj01=random.randint(24,60)
    sj001=round(sj01/10,1)
    print('waiting time:',sj001)
    sleep(sj001)

    if biao01==1:
        print('finish scraping*********')
        break


print('finish the process。。。。。。。。。。。。。')

while 1:
    input('>>>')
